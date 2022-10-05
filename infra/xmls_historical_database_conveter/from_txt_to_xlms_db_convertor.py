from collections import namedtuple
from openpyxl import Workbook


class XlmsConverter:
    file_excel = Workbook()
    plan_a = file_excel.active
    plan_a['A1'] = 'CONCURSO'
    plan_a['B1'] = 'DATA'
    plan_a['C1'] = 'P1'
    plan_a['D1'] = 'P2'
    plan_a['E1'] = 'P3'
    plan_a['F1'] = 'P4'
    plan_a['G1'] = 'P5'
    plan_a['H1'] = 'P6'
    plan_a['I1'] = 'P7'
    plan_a['J1'] = 'P8'
    plan_a['K1'] = 'P9'
    plan_a['L1'] = 'P10'
    plan_a['M1'] = 'P11'
    plan_a['N1'] = 'P12'
    plan_a['O1'] = 'P13'
    plan_a['P1'] = 'P14'
    plan_a['Q1'] = 'P15'
    plan_a['R1'] = 'TOTAL_ARRECADADO'
    plan_a['S1'] = 'GANHADORES_15_ACERTOS'

    def __init__(self):
        self.export_to_xlms_from_txt_result = namedtuple(typename='XlsxConverterFromTxt',
                                                         field_names='last_game, conta_linha, xlms_file_path')

    def export_to_xlms_from_txt(self, txt_file_path, xlsx_folder_path) -> 'XlsxConverterFromTxt':
        """
        Export db of historical games from .TXT to .XLSX
        :param txt_file_path: path of txt file
        :param xlsx_folder_path: path of folder to save the .XLSX file
        :return: attributes values
        """
        try:
            with open(txt_file_path, 'r') as concursos:
                conta_linha = 2
                for linha in concursos:
                    linha = linha.split(',')
                    for coluna in range(1, 20):
                        self.plan_a.cell(row=conta_linha, column=coluna, value=linha[coluna-1])
                    conta_linha += 1
                conta_linha -= 2
                last_game = self.plan_a.cell(row=self.plan_a.max_row, column=1).value
                xlms_file_path = f'{xlsx_folder_path}lotofacil_hist_concursos_{last_game}.xlsx'
                self.file_excel.save(xlms_file_path)
            return self.export_to_xlms_from_txt_result(
                last_game=int(last_game),
                conta_linha=conta_linha,
                xlms_file_path=xlms_file_path
            )
        except Exception as error:
            raise Exception(error.__class__)

    def export_to_xlms_from_sqlite(self):
        pass
